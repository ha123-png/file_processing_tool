import os
import threading
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
import app.shared as shared
from app.shared import (
    get_config_snapshot, OUTPUT_DIR, send_sse, logger,
    THIN_BORDER, HEADER_FILL, HEADER_FONT, CELL_ALIGNMENT, HEADER_ALIGNMENT,
    _excel_sheets, _EXCEL_LOCK,
    _file_row_map, _FILE_ROW_LOCK,
    _file_item_count
)

def _safe_fs_component(name):
    if not name:
        return "sheet"
    invalid = '<>:"/\\|?*\n\r\t'
    s = "".join((ch if ch not in invalid else "_") for ch in str(name)).strip() or "sheet"
    return s[:180]

def _excel_disk_path(sheet_name):
    return OUTPUT_DIR / f"__sheet_{_safe_fs_component(sheet_name)}.xlsx"

def get_excel_state():
    with _EXCEL_LOCK:
        return _excel_sheets.get(shared._excel_active) if shared._excel_active else None

def save_current_sheet():
    with _EXCEL_LOCK:
        if shared._excel_active and shared._excel_active in _excel_sheets:
            _excel_sheets[shared._excel_active]["wb"].save(str(_excel_disk_path(shared._excel_active)))

def init_excel_session(template, name=None):
    if not name:
        name = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_current_sheet()
    wb = Workbook()
    ws = wb.active
    ws.title = "提取结果"
    all_fields = template.get("fields", []) if template else []
    header_fields = [f for f in all_fields if f.get("section") == "header"]
    item_fields = [f for f in all_fields if f.get("section") == "item"]
    all_cols = header_fields + item_fields
    field_labels_list = [f["label"] for f in all_cols] if all_cols else []
    _excel_sheets[name] = {"wb": wb, "ws": ws, "keys": field_labels_list, "labels": field_labels_list, "template": template}
    shared._excel_active = name
    shared._excel_labels = field_labels_list
    if not all_cols:
        return name
    thin_border = THIN_BORDER
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    for ci, f in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=f["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = HEADER_ALIGNMENT
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = 18
    return name

def merge_to_excel(result_data, template, replace=False, file_key=None):
    with _EXCEL_LOCK:
        state = get_excel_state()
        if state is None:
            init_excel_session(template)
            state = get_excel_state()
        ws = state["ws"]
        if replace and ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    all_fields = template.get("fields", []) if template else []
    header_fields = [f for f in all_fields if f.get("section") == "header"]
    item_fields = [f for f in all_fields if f.get("section") == "item"]
    all_cols = header_fields + item_fields
    new_keys = [f["label"] for f in all_cols]
    if state["keys"] and state["keys"] != new_keys:
        tpl_name = (template.get("name") or "新表").strip()
        new_name = tpl_name + " - " + datetime.now().strftime("%H%M%S")
        send_sse("log", {"level": "info", "message": f"模板字段已变更，已自动新建表: {new_name}"})
        with _EXCEL_LOCK:
            save_current_sheet()
            init_excel_session(template, new_name)
            state = get_excel_state()
            ws = state["ws"]
        with _FILE_ROW_LOCK:
            _file_row_map.clear()
    state["keys"] = new_keys
    thin_border = THIN_BORDER
    hdr = result_data.get("header", {})
    items = result_data.get("items", [])
    next_row = ws.max_row + 1
    if not items:
        row_num = next_row
        for ci, f in enumerate(all_cols, 1):
            cell = ws.cell(row=row_num, column=ci, value=hdr.get(f["label"], ""))
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT
        if file_key:
            with _FILE_ROW_LOCK:
                _file_row_map[file_key] = {0: row_num}
                _file_item_count[file_key] = 1
        shared._excel_labels = state["labels"]
        return
    start_row = next_row
    row_numbers = []
    for item in items:
        for ci, f in enumerate(all_cols, 1):
            val = hdr.get(f["label"], "") if f.get("section") == "header" else item.get(f["label"], "")
            cell = ws.cell(row=next_row, column=ci, value=val if val else "")
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT
        row_numbers.append(next_row)
        next_row += 1
    if file_key:
        with _FILE_ROW_LOCK:
            _file_row_map[file_key] = dict(enumerate(row_numbers))
            _file_item_count[file_key] = len(items)
    shared._excel_labels = state["labels"]

def _find_next_data_row(ws):
    max_row = ws.max_row
    if max_row is None:
        return 2
    for r in range(max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                return r + 1
    return 2

def _match_template_excel_columns(ws, template):
    all_fields = template.get("fields", []) if template else []
    all_cols = [f for f in all_fields if f.get("section") in ("header", "item")]
    if not all_cols:
        return None, "模板无可用字段"
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = []
    for v in row:
        if v is None:
            headers.append("")
        else:
            headers.append(str(v).strip())
    while len(headers) < ws.max_column:
        headers.append("")
    col_indices = []
    for f in all_cols:
        label = (f.get("label") or "").strip()
        found = None
        for idx in range(1, ws.max_column + 1):
            h = headers[idx - 1] if idx <= len(headers) else ""
            if h == label:
                found = idx
                break
        col_indices.append(found)
    return (all_cols, col_indices), None

def _workbook_from_mapped_sheet(ws, all_cols, col_indices):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "导入数据"
    thin_border = THIN_BORDER
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    for ci, f in enumerate(all_cols, 1):
        cell = new_ws.cell(row=1, column=ci, value=f["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = HEADER_ALIGNMENT
    for col_cells in new_ws.columns:
        col_letter = col_cells[0].column_letter
        new_ws.column_dimensions[col_letter].width = 18
    data_rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = []
        for src_ci in col_indices:
            if src_ci is None:
                row_vals.append("")
            else:
                row_vals.append(ws.cell(row=r, column=src_ci).value)
        if any(v not in (None, "") for v in row_vals):
            data_rows.append(row_vals)
    for ri, row_vals in enumerate(data_rows, start=2):
        for out_ci, val in enumerate(row_vals, 1):
            cell = new_ws.cell(row=ri, column=out_ci, value=val if val is not None else "")
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT
    field_labels_list = [f["label"] for f in all_cols]
    return new_wb, new_ws, field_labels_list, field_labels_list

def _match_csv_columns(headers, template):
    all_fields = template.get("fields", []) if template else []
    all_cols = [f for f in all_fields if f.get("section") in ("header", "item")]
    if not all_cols:
        return None, "模板无可用字段"
    col_indices = []
    for f in all_cols:
        label = (f.get("label") or "").strip()
        found = None
        for idx, h in enumerate(headers, 1):
            if h == label:
                found = idx
                break
        col_indices.append(found)
    return (all_cols, col_indices), None

def _workbook_from_csv_data(all_cols, col_indices, data_rows):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "导入数据"
    thin_border = THIN_BORDER
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    for ci, f in enumerate(all_cols, 1):
        cell = new_ws.cell(row=1, column=ci, value=f["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = HEADER_ALIGNMENT
    for col_cells in new_ws.columns:
        col_letter = col_cells[0].column_letter
        new_ws.column_dimensions[col_letter].width = 18
    out_rows = []
    for row_vals in data_rows:
        out_row = []
        for src_ci in col_indices:
            if src_ci is None:
                out_row.append("")
            elif src_ci - 1 < len(row_vals):
                out_row.append(row_vals[src_ci - 1])
            else:
                out_row.append("")
        if any(v not in (None, "") for v in out_row):
            out_rows.append(out_row)
    for ri, row_vals in enumerate(out_rows, start=2):
        for out_ci, val in enumerate(row_vals, 1):
            cell = new_ws.cell(row=ri, column=out_ci, value=val if val is not None else "")
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT
    field_labels_list = [f["label"] for f in all_cols]
    return new_wb, new_ws, field_labels_list, field_labels_list

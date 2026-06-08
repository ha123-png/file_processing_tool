import os
import re
import json
import csv
import io
import queue
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from flask import (
    Flask, render_template, request, jsonify,
    Response, send_from_directory, stream_with_context
)

import app.shared as shared

from app.shared import (
    BASE_DIR, CONFIG_PATH, CACHE_DIR,
    get_config_snapshot, save_config, logger,
    TEXT_FORMATS, DOCX_FORMATS, IMAGE_FORMATS, PDF_FORMATS, ALL_SUPPORTED,
    THIN_BORDER, HEADER_FILL, HEADER_FONT, CELL_ALIGNMENT, HEADER_ALIGNMENT,
    task_queue, processing_event, pause_event, abort_event,
    cancelled_task_ids, CANCELLED_TASK_LOCK,
    sse_clients, sse_clients_lock,
    _file_row_map, _FILE_ROW_LOCK,
    _file_item_count,
    _prompt_override, _PROMPT_LOCK,
    get_current_mode, set_current_mode,
    _excel_sheets, _excel_active, _excel_labels, _EXCEL_LOCK,
    send_sse,
    clear_interrupted, get_interrupted
)

from app.processor import get_active_template
from app.file_handler import detect_encoding
from app.excel_manager import (
    get_excel_state, save_current_sheet, init_excel_session,
    merge_to_excel, _find_next_data_row, _safe_fs_component, _excel_disk_path
)
from app.queue_manager import start_worker, get_queue_info, get_queue_info_v2, enqueue_task_id, purge_cancelled_from_queue, cleanup_upload_dir
from app.splitter import split_pdf, split_text, should_split
from app.stats_db import get_dashboard_data, clear_stats

def _mask_api_key(key):
    if not key or len(key) <= 6:
        return key
    return key[:4] + "***" + key[-4:]

def _safe_log_body(key):
    body = json.loads(json.dumps(key))
    llm = body.get("llm", {})
    if llm.get("api_key") and "***" not in llm["api_key"]:
        llm["api_key"] = _mask_api_key(llm["api_key"])
    config_llm = body.get("config", {}).get("llm", {})
    if config_llm.get("api_key") and "***" not in config_llm["api_key"]:
        config_llm["api_key"] = _mask_api_key(config_llm["api_key"])
    return json.dumps(body, ensure_ascii=False)

app = Flask(__name__, template_folder=str(BASE_DIR / "templates"), static_folder=str(BASE_DIR / "static"))
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
app.config["TEMPLATES_AUTO_RELOAD"] = True

import app.task_store as _ts
_ts.init_db()
conn = _ts._get_conn()
try:
    conn.execute("DELETE FROM task_store")
    conn.commit()
finally:
    conn.close()

@app.route("/")
def index():
    return render_template("index.html")

@app.errorhandler(413)
def request_entity_too_large(error):
    return jsonify({"error": "文件过大，最大支持20MB"}), 413

@app.errorhandler(500)
def internal_server_error(error):
    return jsonify({"error": "服务器内部错误，请稍后重试"}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "接口不存在"}), 404

@app.errorhandler(405)
def method_not_allowed(error):
    return jsonify({"error": "请求方法不正确"}), 405

@app.route("/api/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "未检测到上传文件"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400

    original_name = file.filename
    original_name = original_name.replace('\\', '/').split('/')[-1]
    file_size = 0
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)

    max_size = app.config["MAX_CONTENT_LENGTH"]
    if file_size > max_size:
        return jsonify({"error": f"文件过大（{file_size}字节），最大支持{max_size//1024//1024}MB"}), 400

    ext = Path(original_name).suffix.lower()
    if ext not in ALL_SUPPORTED:
        return jsonify({"error": f"不支持的文件格式: {ext}"}), 400

    safe_name = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{original_name}"
    filepath = shared.UPLOAD_DIR / safe_name
    file.save(str(filepath))

    current_mode = request.form.get("mode", "") or get_current_mode()

    # 提取模式不允许大文件（PDF≥5页/文本≥5000字）
    if current_mode == "extract" and should_split(str(filepath), original_name):
        return jsonify({"error": "提取模式不支持大文件（PDF≥5页或文本≥5000字），请使用脱敏模式处理"}), 400

    if task_queue.qsize() >= 100:
        return jsonify({"error": "队列已满（上限100），请等待当前任务完成后再添加"}), 429

    purge_cancelled_from_queue()

    from app.shared import TaskState, QueueMode as QMode, TaskStatus as TStatus
    from app.task_store import save as ts_save
    import uuid as _uuid
    parent_task_id = str(_uuid.uuid4())
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    qmode = QMode.EXTRACT if current_mode == "extract" else QMode.DESENSITIZE
    task_id = str(_uuid.uuid4())

    is_large = current_mode == "desensitize" and should_split(str(filepath), original_name)

    if is_large:
        parent = TaskState(
            task_id=parent_task_id, display_name=original_name,
            original_name=original_name, filepath=str(filepath),
            mode=qmode, status=TStatus.WAITING,
            created_at=now_str,
        )
        ts_save(parent)

        ext_lower = Path(original_name).suffix.lower()
        if ext_lower == '.pdf':
            chunks = split_pdf(str(filepath), original_name, shared.UPLOAD_DIR, parent_task_id)
        else:
            chunks = split_text(str(filepath), original_name, shared.UPLOAD_DIR, parent_task_id)

        if not chunks:
            ts_save(TaskState(task_id=task_id, display_name=original_name,
                original_name=original_name, filepath=str(filepath),
                mode=qmode, status=TStatus.WAITING, created_at=now_str))
            task_queue.put(task_id)
        else:
            chunk_ids = []
            for c in chunks:
                c_name = c["display_name"]
                chunk_ext = os.path.splitext(c["filepath"])[1].lower()
                cid = str(_uuid.uuid4())
                ts_save(TaskState(task_id=cid, display_name=c_name,
                    original_name=c_name, filepath=c["filepath"],
                    mode=qmode, status=TStatus.WAITING,
                    is_chunk=True, parent_task_id=parent_task_id,
                    chunk_index=c["chunk_index"], total_chunks=c["total_chunks"],
                    file_ext=chunk_ext,
                    created_at=now_str))
                chunk_ids.append(cid)
            for cid in chunk_ids:
                task_queue.put(cid)
            task_id = parent_task_id

        send_sse("log", {"level": "info", "message": f"[{original_name}] 大文件已自动拆分为 {len(chunks)} 个子任务"})
        return jsonify({
            "task_id": task_id, "mode": current_mode,
            "is_large": True, "total_chunks": len(chunks) if chunks else 1,
            "queue_size": task_queue.qsize()
        })
    else:
        ts_save(TaskState(task_id=task_id, display_name=original_name,
            original_name=original_name, filepath=str(filepath),
            mode=qmode, status=TStatus.WAITING, created_at=now_str))
        task_queue.put(task_id)

    qsize = task_queue.qsize()
    send_sse("log", {"level": "info", "message": f"[{original_name}] 已加入队列（第{qsize}位）"})
    return jsonify({"message": "文件已加入队列", "file": original_name, "queue_size": qsize, "task_id": task_id, "mode": current_mode})

@app.route("/api/events")
def sse_events():
    def generate():
        client_queue = queue.Queue(maxsize=200)
        with sse_clients_lock:
            sse_clients.append(client_queue)
        try:
            while not task_queue.empty():
                try:
                    task_queue.get_nowait()
                except:
                    break
            abort_event.set()
            processing_event.clear()
            pause_event.clear()
            clear_interrupted()

            qinfo = get_queue_info()
            qinfo_v2 = get_queue_info_v2()
            qinfo["v2"] = qinfo_v2
            yield f"event: connected\ndata: {json.dumps(qinfo, ensure_ascii=False)}\n\n"
            while True:
                try:
                    msg = client_queue.get(timeout=30)
                    yield msg
                except queue.Empty:
                    yield ": keepalive\n\n"
        except GeneratorExit:
            with sse_clients_lock:
                if client_queue in sse_clients:
                    sse_clients.remove(client_queue)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )

@app.route("/api/status")
def api_status():
    return jsonify(get_queue_info())

@app.route("/api/stats/dashboard")
def api_stats_dashboard():
    mode = request.args.get("mode")
    if mode and mode not in ("desensitize", "extract"):
        mode = None
    return jsonify(get_dashboard_data(mode))

@app.route("/api/stats/clear", methods=["POST"])
def api_stats_clear():
    count = clear_stats()
    logger.info(f"统计数据已清除: {count} 条记录")
    return jsonify({"message": f"已清除 {count} 条统计记录"})

@app.route("/api/kill_queue", methods=["POST"])
def api_kill_queue():
    body = request.get_json(force=True) or {}
    task_ids = body.get("task_ids", [])
    if not task_ids:
        return jsonify({"error": "未指定要移除的任务"}), 400
    with CANCELLED_TASK_LOCK:
        for tid in task_ids:
            cancelled_task_ids.add(tid)
    logger.warning(f"从队列中移除 {len(task_ids)} 个任务")
    send_sse("log", {"level": "warning", "message": f"已从队列中移除 {len(task_ids)} 个任务"})
    return jsonify({"message": f"已标记 {len(task_ids)} 个任务为取消", "cancelled": len(task_ids)})

@app.route("/api/download/<filename>")
def download_file(filename):
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({"error": "不允许的文件名"}), 400
    if filename.startswith("__orig_"):
        cached = CACHE_DIR / filename
        if cached.exists():
            return send_from_directory(CACHE_DIR, filename, as_attachment=True)
    return send_from_directory(shared.OUTPUT_DIR, filename, as_attachment=True)

@app.route("/api/outputs")
def list_outputs():
    files = []
    for f in sorted(shared.OUTPUT_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        if f.is_file():
            files.append({
                "name": f.name,
                "size": f.stat().st_size,
                "mtime": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    return jsonify(files)

@app.route("/api/preview/<filename>")
def preview_file(filename):
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({"error": "不允许的文件名"}), 400
    filepath = shared.OUTPUT_DIR / filename
    if not filepath.exists():
        return jsonify({"error": "文件不存在"}), 404
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    return jsonify({"name": filename, "content": content[:50000]})

@app.route("/api/save", methods=["POST"])
def save_file():
    body = request.get_json(force=True)
    if not body or "filename" not in body or "content" not in body:
        return jsonify({"error": "缺少参数 filename 或 content"}), 400
    filename = body["filename"]
    if ".." in filename or "/" in filename or "\\" in filename:
        return jsonify({"error": "非法的文件名"}), 400
    filepath = shared.OUTPUT_DIR / filename
    if not filepath.exists():
        return jsonify({"error": "文件不存在"}), 404
    content = body["content"]
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    logger.info(f"文件已保存: {filename} ({len(content)}字符)")
    return jsonify({"message": "保存成功", "filename": filename, "size": len(content)})

@app.route("/api/prompt", methods=["GET", "POST"])
def api_prompt():
    if request.method == "POST":
        body = request.get_json(force=True)
        if not body:
            return jsonify({"error": "无效的提示词数据"}), 400
        allowed = {"first_pass", "second_pass"}
        for key in body:
            if key not in allowed:
                return jsonify({"error": f"不允许修改: {key}"}), 400
        with _PROMPT_LOCK:
            _prompt_override.clear()
            for k, v in body.items():
                if v and v.strip():
                    _prompt_override[k] = v.strip()
        with _PROMPT_LOCK:
            ov = dict(_prompt_override)
        logger.info(f"提示词覆盖已更新")
        return jsonify({"message": "提示词已更新（仅当前会话有效）", "overrides": ov})
    cfg = get_config_snapshot()
    with _PROMPT_LOCK:
        ov = dict(_prompt_override)
    return jsonify({
        "overrides": ov,
        "defaults": {
            "first_pass": cfg.get("prompt", {}).get("first_pass", ""),
            "second_pass": cfg.get("prompt", {}).get("second_pass", "")
        }
    })

@app.route("/api/merge_extraction", methods=["POST"])
def api_merge_extraction():
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "无效数据"}), 400
    template = get_active_template()
    if not template:
        return jsonify({"error": "未配置提取模板"}), 400
    result_data = body.get("result", {})
    replace = body.get("replace", False)
    merge_to_excel(result_data, template, replace=replace)
    with _EXCEL_LOCK:
        state = _excel_sheets.get(shared._excel_active) if shared._excel_active else None
        row_count = state["ws"].max_row - 1 if state else 0
        name = shared._excel_active
    return jsonify({"message": "已合并到Excel", "row_count": max(0, row_count), "sheet_name": name})

@app.route("/api/update_file_rows", methods=["POST"])
def api_update_file_rows():
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "无效数据"}), 400
    filename = body.get("filename", "")
    if not filename:
        return jsonify({"error": "缺少 filename 参数"}), 400
    header = body.get("header", {})
    items = body.get("items", [])
    template = get_active_template()
    if not template:
        return jsonify({"error": "未配置提取模板"}), 400
    thin_border = THIN_BORDER
    all_fields = template.get("fields", []) if template else []
    header_fields = [f for f in all_fields if f.get("section") == "header"]
    item_fields = [f for f in all_fields if f.get("section") == "item"]
    all_cols = header_fields + item_fields
    all_col_keys = [f["label"] for f in all_cols]

    def write_item_row(ws, row_num, item, header_vals):
        for ci, f in enumerate(all_cols, 1):
            val = header_vals.get(f["label"], "") if f.get("section") == "header" else item.get(f["label"], "")
            cell = ws.cell(row=row_num, column=ci, value=val if val else "")
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT

    with _EXCEL_LOCK:
        state = get_excel_state()
        if state is None:
            init_excel_session(template)
            state = get_excel_state()
        ws = state["ws"]

        with _FILE_ROW_LOCK:
            old_map = _file_row_map.get(filename, {})

        if not old_map:
            table_keys = state.get("keys", [])
            col_map = {}
            for ci, tk in enumerate(table_keys):
                if tk in all_col_keys:
                    col_map[tk] = ci
            if not col_map:
                for ci, tk in enumerate(table_keys):
                    col_map[tk] = ci
            mapped_cols = [(tk, col_map.get(tk)) for tk in table_keys]
            next_row = _find_next_data_row(ws)
            new_map = {}
            effective_items = items if items else [{}]
            for orig_idx, item in enumerate(effective_items):
                for ci, (tk, src_ci) in enumerate(mapped_cols, 1):
                    if tk in all_col_keys:
                        val = header.get(tk, "") if tk in [h["label"] for h in header_fields] else item.get(tk, "")
                    else:
                        val = ""
                    cell = ws.cell(row=next_row, column=ci, value=val if val else "")
                    cell.border = thin_border
                    cell.alignment = CELL_ALIGNMENT
                new_map[orig_idx] = next_row
                next_row += 1
            with _FILE_ROW_LOCK:
                _file_row_map[filename] = new_map
            save_current_sheet()
            return jsonify({"message": f"文件 '{filename}' 已追加 {len(new_map)} 行"})

        new_map = {}
        total_original = _file_item_count.get(filename, len(items))
        updated = 0
        appended = 0

        for orig_idx in range(len(items)):
            item = items[orig_idx]
            if orig_idx in old_map:
                row_num = old_map[orig_idx]
                write_item_row(ws, row_num, item, header)
                new_map[orig_idx] = row_num
                updated += 1
            elif orig_idx < total_original:
                pass
            else:
                next_row = _find_next_data_row(ws)
                write_item_row(ws, next_row, item, header)
                new_map[orig_idx] = next_row
                appended += 1

        with _FILE_ROW_LOCK:
            _file_row_map[filename] = new_map
            _file_item_count[filename] = max(total_original, len(items))
        save_current_sheet()
        msg = f"已更新 {updated} 行" + (f"，追加 {appended} 行" if appended else "") + f"（文件: {filename}）"
        return jsonify({"message": msg})

@app.route("/api/clear_excel", methods=["POST"])
def api_clear_excel():
    with _EXCEL_LOCK:
        shared._excel_sheets.clear()
        shared._excel_active = None
        shared._excel_labels = None
    for f in shared.OUTPUT_DIR.glob("__sheet_*.xlsx"):
        try:
            f.unlink()
        except OSError:
            pass
    return jsonify({"message": "ok"})

@app.route("/api/export_excel", methods=["POST"])
def api_export_excel():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据可导出"}), 400
    body = request.get_json(force=True) or {}
    raw = (body.get("filename") or "").strip()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if raw:
        if ".." in raw or "/" in raw or "\\" in raw:
            return jsonify({"error": "非法的文件名"}), 400
        base = raw[:-5] if raw.lower().endswith(".xlsx") else raw
        filename = _safe_fs_component(base) + ".xlsx"
    else:
        filename = f"提取结果_{timestamp}.xlsx"
    filepath = shared.OUTPUT_DIR / filename
    with _EXCEL_LOCK:
        state["wb"].save(str(filepath))
    return jsonify({"message": "导出成功", "filename": filename})

@app.route("/api/export_filtered_excel", methods=["POST"])
def api_export_filtered_excel():
    body = request.get_json(force=True)
    if not body or "labels" not in body or "rows" not in body:
        return jsonify({"error": "缺少 labels 或 rows 参数"}), 400
    labels = body["labels"]
    rows = body["rows"]
    sheet_name = body.get("sheet_name", "筛选结果")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    thin_border = THIN_BORDER
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    for ci, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = HEADER_ALIGNMENT
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            cell.border = thin_border
            cell.alignment = CELL_ALIGNMENT
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = 18
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"筛选结果_{timestamp}.xlsx"
    filepath = shared.OUTPUT_DIR / filename
    wb.save(str(filepath))
    return jsonify({"message": "导出成功", "filename": filename})

@app.route("/api/excel_add_row", methods=["POST"])
def api_excel_add_row():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据"}), 400
    ws = state["ws"]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value="")
    row_count = ws.max_row - 1
    return jsonify({"message": "已新增行", "row_count": row_count, "new_row_index": next_row})

@app.route("/api/new_excel_session", methods=["POST"])
def api_new_excel_session():
    template = get_active_template()
    if not template:
        return jsonify({"error": "未配置模板"}), 400
    body = request.get_json(force=True) or {}
    name = body.get("name", "")
    if not name:
        name = datetime.now().strftime("%Y%m%d_%H%M%S")
    init_excel_session(template, name)
    with _EXCEL_LOCK:
        sheets = list(_excel_sheets.keys())
        active = shared._excel_active
    return jsonify({"message": f"已创建新表: {name}", "sheet_name": name, "sheets": sheets, "active": active})

@app.route("/api/list_sheets", methods=["GET"])
def api_list_sheets():
    with _EXCEL_LOCK:
        sheets = list(_excel_sheets.keys())
        active = shared._excel_active
    return jsonify({"sheets": sheets, "active": active})

@app.route("/api/switch_sheet", methods=["POST"])
def api_switch_sheet():
    body = request.get_json(force=True)
    name = body.get("name", "")
    with _EXCEL_LOCK:
        if name not in _excel_sheets:
            return jsonify({"error": "表不存在"}), 400
        save_current_sheet()
        shared._excel_active = name
        state = _excel_sheets[name]
        shared._excel_labels = state.get("labels", [])
    return jsonify({"message": f"已切换到: {name}", "active": name})

@app.route("/api/rename_sheet", methods=["POST"])
def api_rename_sheet():
    body = request.get_json(force=True)
    old = body.get("old_name", "")
    new = body.get("new_name", "")
    with _EXCEL_LOCK:
        if old not in _excel_sheets:
            return jsonify({"error": "表不存在"}), 400
        if new in _excel_sheets:
            return jsonify({"error": "表名已存在"}), 400
        _excel_sheets[new] = _excel_sheets.pop(old)
        if shared._excel_active == old:
            shared._excel_active = new
        old_path = _excel_disk_path(old)
        new_path = _excel_disk_path(new)
        if old_path.exists():
            if new_path.exists() and old_path.resolve() != new_path.resolve():
                new_path.unlink()
            old_path.rename(new_path)
        sheets = list(_excel_sheets.keys())
        active = shared._excel_active
    return jsonify({"message": f"已重命名: {old} -> {new}", "sheets": sheets, "active": active})

@app.route("/api/preview_xlsx_sheets", methods=["POST"])
def api_preview_xlsx_sheets():
    if "file" not in request.files:
        return jsonify({"error": "未检测到文件"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "仅支持 .xlsx 格式"}), 400
    safe_upload = shared.UPLOAD_DIR / f"sheetcheck_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.xlsx"
    file.save(str(safe_upload))
    try:
        wb = load_workbook(safe_upload, data_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
        return jsonify({"sheets": sheet_names, "multi_sheet": len(sheet_names) > 1})
    except Exception as e:
        return jsonify({"error": f"读取文件失败: {str(e)}"}), 400
    finally:
        try:
            if safe_upload.exists():
                safe_upload.unlink()
        except OSError:
            pass

@app.route("/api/delete_sheet", methods=["POST"])
def api_delete_sheet():
    body = request.get_json(force=True)
    name = body.get("name", "")
    with _EXCEL_LOCK:
        if name not in _excel_sheets:
            return jsonify({"error": "表不存在"}), 400
        if name in _excel_sheets:
            try:
                _excel_sheets[name]["wb"].close()
            except Exception:
                pass
            del _excel_sheets[name]
        disk_path = _excel_disk_path(name)
        if disk_path.exists():
            try:
                disk_path.unlink()
            except OSError:
                pass
        if shared._excel_active == name:
            shared._excel_active = next(iter(_excel_sheets)) if _excel_sheets else None
            if shared._excel_active:
                shared._excel_labels = _excel_sheets[shared._excel_active].get("labels", [])
        save_current_sheet()
    return jsonify({"message": f"已删除表: {name}", "active": shared._excel_active})

@app.route("/api/preview_excel", methods=["GET"])
def api_preview_excel():
    state = get_excel_state()
    if state is None:
        return jsonify({"rows": [], "field_keys": [], "field_labels": []})
    ws = state["ws"]
    page = request.args.get("page", type=int)
    page_size = request.args.get("page_size", type=int)
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    total_rows = len(rows) - 1 if rows else 0
    if page and page_size:
        total_pages = max(1, -(-total_rows // page_size)) if total_rows > 0 else 1
        start = (page - 1) * page_size + 1
        end = start + page_size
        page_rows = [rows[0]] + rows[start:end] if len(rows) > 1 else rows
        with _EXCEL_LOCK:
            keys = (state.get("keys") or [])[:]
            labels = (state.get("labels") or [])[:]
        return jsonify({
            "rows": page_rows, "field_keys": keys, "field_labels": labels,
            "total_rows": total_rows, "total_pages": total_pages,
            "page": page, "page_size": page_size
        })
    with _EXCEL_LOCK:
        keys = (state.get("keys") or [])[:]
        labels = (state.get("labels") or [])[:]
    return jsonify({"rows": rows, "field_keys": keys, "field_labels": labels, "total_rows": total_rows})

@app.route("/api/excel_delete_row", methods=["POST"])
def api_excel_delete_row():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据"}), 400
    body = request.get_json(force=True)
    if not body or "row_index" not in body:
        return jsonify({"error": "缺少 row_index"}), 400
    row_idx = int(body["row_index"])
    ws = state["ws"]
    if row_idx < 1 or row_idx >= ws.max_row:
        return jsonify({"error": "行索引无效"}), 400
    ws.delete_rows(row_idx + 1, 1)
    row_count = ws.max_row - 1
    return jsonify({"message": "已删除", "row_count": row_count})

@app.route("/api/excel_batch_delete_rows", methods=["POST"])
def api_excel_batch_delete_rows():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据"}), 400
    body = request.get_json(force=True)
    if not body or "row_indices" not in body:
        return jsonify({"error": "缺少 row_indices"}), 400
    indices = sorted(set(int(x) for x in body["row_indices"]), reverse=True)
    deleted_rows = set(idx + 1 for idx in indices)
    ws = state["ws"]
    for idx in indices:
        actual_row = idx + 1
        if 1 <= idx < ws.max_row:
            ws.delete_rows(actual_row, 1)
    with _FILE_ROW_LOCK:
        for filename, old_map in list(_file_row_map.items()):
            dead = set()
            for item_idx, row in list(old_map.items()):
                if row in deleted_rows:
                    dead.add(item_idx)
            for item_idx in dead:
                del old_map[item_idx]
            for item_idx, row in list(old_map.items()):
                shift = sum(1 for d in deleted_rows if d < row)
                if shift:
                    old_map[item_idx] = row - shift
            if old_map:
                pass
            else:
                _file_row_map.pop(filename, None)
    row_count = ws.max_row - 1
    save_current_sheet()
    return jsonify({"message": f"已删除 {len(indices)} 行", "row_count": row_count})

@app.route("/api/excel_update_cell", methods=["POST"])
def api_excel_update_cell():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据"}), 400
    body = request.get_json(force=True)
    if not body or "row_index" not in body or "col_index" not in body or "value" not in body:
        return jsonify({"error": "缺少参数"}), 400
    ws = state["ws"]
    row_idx = int(body["row_index"])
    col_idx = int(body["col_index"])
    if row_idx < 1 or row_idx >= ws.max_row:
        return jsonify({"error": "行索引无效"}), 400
    ws.cell(row=row_idx + 1, column=col_idx, value=body["value"])
    return jsonify({"message": "已更新"})

@app.route("/api/excel_batch_update", methods=["POST"])
def api_excel_batch_update():
    state = get_excel_state()
    if state is None:
        return jsonify({"error": "没有数据"}), 400
    body = request.get_json(force=True)
    if not body or "cells" not in body:
        return jsonify({"error": "缺少 cells 参数"}), 400
    cells = body["cells"]
    if not isinstance(cells, list):
        return jsonify({"error": "cells 必须是数组"}), 400
    with _EXCEL_LOCK:
        ws = state["ws"]
        for cell in cells:
            row_idx = int(cell.get("row_index", 0))
            col_idx = int(cell.get("col_index", 0))
            value = cell.get("value", "")
            if row_idx < 1 or row_idx >= ws.max_row:
                continue
            if col_idx < 1:
                continue
            ws.cell(row=row_idx + 1, column=col_idx, value=value)
    return jsonify({"message": f"已更新 {len(cells)} 个单元格"})

@app.route("/api/import_excel", methods=["POST"])
def api_import_excel():
    if "file" not in request.files:
        return jsonify({"error": "未检测到上传文件"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400
    if not file.filename.lower().endswith((".xlsx", ".csv")):
        return jsonify({"error": "仅支持 .xlsx 或 .csv 格式"}), 400
    template = get_active_template()
    if not template:
        return jsonify({"error": "未配置提取模板"}), 400
    is_csv = file.filename.lower().endswith(".csv")
    ext = ".csv" if is_csv else ".xlsx"
    safe_upload = shared.UPLOAD_DIR / f"import_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
    file.save(str(safe_upload))
    wb = None
    multi_sheet = False
    available_sheets = []
    try:
        headers = []
        data_rows = []
        if is_csv:
            encoding = detect_encoding(str(safe_upload))
            with open(safe_upload, "r", encoding=encoding) as f:
                content = f.read()
            if content.startswith('\ufeff'):
                content = content[1:]
            reader = csv.reader(io.StringIO(content))
            csv_rows = list(reader)
            if not csv_rows:
                return jsonify({"error": "CSV文件为空"}), 400
            headers = [h.strip() for h in csv_rows[0]]
            data_rows = csv_rows[1:]
        else:
            wb = load_workbook(safe_upload, data_only=True)
            requested_sheet = request.args.get("sheet", "")
            available_sheets = list(wb.sheetnames)
            multi_sheet = len(available_sheets) > 1
            if requested_sheet and requested_sheet in available_sheets:
                ws = wb[requested_sheet]
            elif multi_sheet:
                ws = wb[available_sheets[0]]
            else:
                ws = wb.active
            merged_list = list(ws.merged_cells.ranges)
            if merged_list:
                return jsonify({"error": "表头或数据区域含合并单元格，请取消合并后重试"}), 400
            for cell in ws[1]:
                headers.append(str(cell.value).strip() if cell.value is not None else "")
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append([str(v) if v is not None else "" for v in row])
        data_rows = [row for row in data_rows if any(v.strip() for v in row)]
        if not data_rows:
            return jsonify({"error": "文件无有效数据行"}), 400
        if len(data_rows) > 500:
            return jsonify({"error": f"单次最多导入500行，当前文件含{len(data_rows)}行数据，请拆分后分批导入"}), 400
        thin_border = THIN_BORDER
        with _EXCEL_LOCK:
            existing_state = get_excel_state()
            if existing_state is not None and shared._excel_active:
                ws = existing_state["ws"]
                current_rows = ws.max_row - 1
                if current_rows + len(data_rows) > 800:
                    return jsonify({"error": f"单表最多800行，当前已有{current_rows}行，本次导入{len(data_rows)}行将超限，请新建表后导入", "suggest_new_sheet": True}), 400
                existing_keys = existing_state.get("keys", [])
                existing_labels = existing_state.get("labels", [])
                file_header_map = {}
                for idx, h in enumerate(headers, 1):
                    file_header_map[h] = idx
                matched_file_cols = set()
                source_indices = []
                new_cols_added = 0
                for k, l in zip(existing_keys, existing_labels):
                    found = None
                    for h, col_idx in file_header_map.items():
                        if h == l or (k and h == k):
                            found = col_idx
                            matched_file_cols.add(col_idx)
                            break
                    source_indices.append(found)
                extra_file_indices = [i for i in range(1, len(headers) + 1) if i not in matched_file_cols]
                extra_labels = [headers[i - 1] for i in extra_file_indices]
                if extra_labels:
                    new_cols_added = len(extra_labels)
                    all_indices = source_indices + extra_file_indices
                    new_header_keys = existing_keys + extra_labels
                    new_header_labels = existing_labels + extra_labels
                    existing_state["keys"] = new_header_keys
                    existing_state["labels"] = new_header_labels
                    for i, lbl in enumerate(extra_labels):
                        col_pos = len(existing_keys) + i + 1
                        cell = ws.cell(row=1, column=col_pos, value=lbl)
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.border = thin_border
                        cell.alignment = HEADER_ALIGNMENT
                else:
                    all_indices = source_indices
                appended = 0
                n_cols = len(all_indices)
                for row_vals in data_rows:
                    row_num = ws.max_row + 1
                    for ci in range(1, n_cols + 1):
                        src_idx = all_indices[ci - 1]
                        if src_idx is not None and src_idx - 1 < len(row_vals):
                            val = row_vals[src_idx - 1]
                        else:
                            val = ""
                        cell = ws.cell(row=row_num, column=ci, value=val if val else "")
                        cell.border = thin_border
                        cell.alignment = CELL_ALIGNMENT
                    appended += 1
                save_current_sheet()
                msg = f"已追加 {appended} 行到当前表"
                if new_cols_added:
                    msg += f"，自动补充 {new_cols_added} 个新列"
                resp = {"message": msg, "sheet_name": shared._excel_active, "appended": appended}
                if multi_sheet:
                    resp["import_sheets"] = available_sheets
                    resp["multi_sheet"] = True
                return jsonify(resp)
            template_fields = template.get("fields", []) if template else []
            template_cols = [f for f in template_fields if f.get("section") in ("header", "item")]
            if len(data_rows) > 800:
                return jsonify({"error": f"单表最多800行，当前文件含{len(data_rows)}行数据，请拆分后分批导入或使用更小的文件"}), 400
            template_source_indices = []
            for f in template_cols:
                label = (f.get("label") or "").strip()
                found = None
                for idx, h in enumerate(headers, 1):
                    if h == label:
                        found = idx
                        break
                template_source_indices.append(found)
            matched_header_set = set()
            for idx in template_source_indices:
                if idx is not None:
                    matched_header_set.add(idx)
            extra_indices = [i for i in range(1, len(headers) + 1) if i not in matched_header_set]
            all_cols = list(template_cols)
            for ei in extra_indices:
                all_cols.append({"label": headers[ei - 1], "section": "item"})
            all_source_indices = list(template_source_indices) + extra_indices
            field_labels_list = [f["label"] for f in all_cols]
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "提取结果"
            header_fill = HEADER_FILL
            header_font = HEADER_FONT
            for ci, f in enumerate(all_cols, 1):
                cell = new_ws.cell(row=1, column=ci, value=f["label"])
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = HEADER_ALIGNMENT
            for col_cells in new_ws.columns:
                col_letter = col_cells[0].column_letter
                new_ws.column_dimensions[col_letter].width = 18
            for row_vals in data_rows:
                row_num = new_ws.max_row + 1
                for ci, src_idx in enumerate(all_source_indices, 1):
                    if src_idx is not None and src_idx - 1 < len(row_vals):
                        val = row_vals[src_idx - 1]
                    else:
                        val = ""
                    cell = new_ws.cell(row=row_num, column=ci, value=val if val else "")
                    cell.border = thin_border
                    cell.alignment = CELL_ALIGNMENT
            name = (request.form.get("name") or "").strip() or datetime.now().strftime("%Y%m%d_%H%M%S")
            if name in _excel_sheets:
                return jsonify({"error": "表名已存在，请更换名称"}), 400
            save_current_sheet()
            _excel_sheets[name] = {
                "wb": new_wb,
                "ws": new_ws,
                "keys": field_labels_list,
                "labels": field_labels_list,
                "template": template
            }
            shared._excel_active = name
            shared._excel_labels = field_labels_list
            new_wb.save(str(_excel_disk_path(name)))
            sheets = list(_excel_sheets.keys())
            active = shared._excel_active
        resp = {"message": "导入成功", "sheet_name": name, "sheets": sheets, "active": active}
        if multi_sheet:
            resp["import_sheets"] = available_sheets
            resp["multi_sheet"] = True
        return jsonify(resp)
    except Exception as e:
        logger.exception("import_excel")
        return jsonify({"error": f"导入失败: {str(e)}"}), 400
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            if safe_upload.exists():
                safe_upload.unlink()
        except OSError:
            pass

@app.route("/api/save_settings", methods=["POST"])
def api_save_settings():
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "无效数据"}), 400
    from app.shared import CONFIG_LOCK, config

    need_restart = False
    errors = []

    with CONFIG_LOCK:
        prompt_data = body.get("prompt", {})
        if prompt_data:
            for k, v in prompt_data.items():
                if k not in ("first_pass", "second_pass"):
                    errors.append(f"不允许修改提示词: {k}")
            if not errors:
                with _PROMPT_LOCK:
                    _prompt_override.clear()
                    for k, v in prompt_data.items():
                        if v and v.strip():
                            _prompt_override[k] = v.strip()

        cfg_data = body.get("config", {})
        templates_data = body.get("templates")

        if cfg_data or templates_data is not None:
            new_config = json.loads(json.dumps(config))
            allowed_keys = {"llm", "lm_studio", "desensitization", "server", "extraction"}
            for key in list(cfg_data.keys()):
                if key not in allowed_keys:
                    errors.append(f"不允许修改配置项: {key}")
                    del cfg_data[key]
            if "server" in cfg_data:
                need_restart = True
            for key, val in cfg_data.items():
                if isinstance(val, dict):
                    if key not in new_config:
                        new_config[key] = {}
                    new_config[key].update(val)
                else:
                    new_config[key] = val
            if templates_data is not None:
                if "extraction" not in new_config:
                    new_config["extraction"] = {}
                new_config["extraction"]["templates"] = templates_data
            if "llm" in cfg_data and "***" in cfg_data["llm"].get("api_key", ""):
                old_llm = config.get("llm", {})
                if old_llm.get("api_key"):
                    new_config["llm"]["api_key"] = old_llm["api_key"]
            if not errors:
                save_config(new_config)

    if errors:
        return jsonify({"error": "; ".join(errors)}), 400

    msg = "设置已保存"
    if need_restart:
        msg += "，端口修改需重启服务生效"
    return jsonify({"message": msg, "restart_required": need_restart})

@app.route("/api/mode", methods=["GET", "POST"])
def api_mode():
    if request.method == "POST":
        body = request.get_json(force=True)
        if not body or "mode" not in body:
            return jsonify({"error": "缺少 mode 参数"}), 400
        m = body["mode"]
        if m not in ("desensitize", "extract"):
            return jsonify({"error": "mode 必须为 desensitize 或 extract"}), 400
        set_current_mode(m)
        logger.info(f"工作模式已切换: {m}")
        return jsonify({"mode": m, "message": f"已切换到{'脱敏' if m == 'desensitize' else '提取'}模式"})
    return jsonify({"mode": get_current_mode()})

@app.route("/api/templates", methods=["GET", "POST"])
def api_templates():
    if request.method == "POST":
        body = request.get_json(force=True)
        if not body or "templates" not in body:
            return jsonify({"error": "缺少 templates 参数"}), 400
        from app.shared import CONFIG_LOCK, config
        with CONFIG_LOCK:
            new_config = json.loads(json.dumps(config))
            if "extraction" not in new_config:
                new_config["extraction"] = {}
            new_config["extraction"]["templates"] = body["templates"]
            save_config(new_config)
        logger.info(f"提取模板已更新: {len(body['templates'])} 个模板")
        tpl = get_config_snapshot().get("extraction", {}).get("templates", [])
        return jsonify({"message": "模板已更新", "templates": tpl})
    tpl = get_config_snapshot().get("extraction", {}).get("templates", [])
    return jsonify({"templates": tpl})

@app.route("/api/validate_invoice", methods=["POST"])
def api_validate_invoice():
    body = request.get_json(force=True)
    if not body:
        return jsonify({"error": "无效数据"}), 400
    items = body.get("items", [])
    header = body.get("header", {})
    tolerance = body.get("tolerance", None)
    if tolerance is None:
        tolerance = get_config_snapshot().get("extraction", {}).get("invoice_tolerance", 0.02)
    else:
        try:
            tolerance = float(tolerance)
        except (ValueError, TypeError):
            return jsonify({"error": "容差参数无效"}), 400
    from app.invoice_validator import validate_invoice_items
    rules = body.get("rules", None)
    result = validate_invoice_items(items, header, tolerance, rules)
    return jsonify(result)

@app.route("/api/llm_profiles", methods=["GET", "POST"])
def api_llm_profiles():
    from app.shared import CONFIG_LOCK, config
    if request.method == "POST":
        body = request.get_json(force=True)
        if not body or "profiles" not in body:
            return jsonify({"error": "缺少 profiles 参数"}), 400
        profiles = body["profiles"]
        if not isinstance(profiles, list):
            return jsonify({"error": "profiles 必须是数组"}), 400
        for p in profiles:
            if not isinstance(p, dict) or "name" not in p:
                return jsonify({"error": "每个方案必须包含 name 字段"}), 400
            if p.get("api_key") and "***" in p["api_key"]:
                old_profiles = config.get("llm_profiles", {}).get("profiles", [])
                for op in old_profiles:
                    if op.get("name") == p["name"] and op.get("api_key"):
                        p["api_key"] = op["api_key"]
                        break
        with CONFIG_LOCK:
            new_config = json.loads(json.dumps(config))
            if "llm_profiles" not in new_config:
                new_config["llm_profiles"] = {}
            new_config["llm_profiles"]["profiles"] = profiles
            save_config(new_config)
        logger.info(f"LLM配置方案已更新: {len(profiles)} 个方案")
        return jsonify({"profiles": profiles, "message": f"已保存 {len(profiles)} 个方案"})
    profiles = get_config_snapshot().get("llm_profiles", {}).get("profiles", [])
    safe = []
    for p in profiles:
        pc = dict(p)
        if pc.get("api_key") and "***" not in pc.get("api_key", ""):
            pc["api_key"] = _mask_api_key(pc["api_key"])
        safe.append(pc)
    return jsonify({"profiles": safe})

@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    if request.method == "POST":
        body = request.get_json(force=True)
        if not body:
            return jsonify({"error": "无效的配置数据"}), 400
        from app.shared import CONFIG_LOCK, config
        with CONFIG_LOCK:
            allowed_keys = {"llm", "lm_studio", "desensitization", "server", "extraction"}
            for key in body:
                if key not in allowed_keys:
                    return jsonify({"error": f"不允许修改配置项: {key}"}), 400
            new_config = json.loads(json.dumps(config))
            for key, val in body.items():
                if isinstance(val, dict):
                    if key not in new_config:
                        new_config[key] = {}
                    new_config[key].update(val)
                else:
                    new_config[key] = val
            if "llm" in body and "***" in body["llm"].get("api_key", ""):
                old_llm = config.get("llm", {})
                if old_llm.get("api_key"):
                    new_config["llm"]["api_key"] = old_llm["api_key"]
            save_config(new_config)
        logger.info(f"配置已更新: {_safe_log_body(body)}")
        need_restart = "server" in body
        msg = "配置已更新"
        if need_restart:
            msg += "，端口修改需重启服务生效"
        return jsonify({"message": msg, "restart_required": need_restart, "config": config})
    cfg = get_config_snapshot()
    llm_cfg = cfg.get("llm", {})
    if not llm_cfg.get("base_url"):
        legacy = cfg.get("lm_studio", {})
        llm_cfg = {
            "provider": "lm_studio",
            "base_url": f"http://{legacy.get('host', '127.0.0.1')}:{legacy.get('port', '1234')}/v1",
            "api_key": legacy.get("api_key", ""),
            "model": legacy.get("model", ""),
            "temperature": legacy.get("temperature", 0.3),
            "max_tokens": legacy.get("max_tokens", 4096),
            "timeout": legacy.get("timeout", 300),
            "multimodal": True,
            "enable_thinking": legacy.get("enable_thinking", True),
            "reasoning_effort": legacy.get("reasoning_effort")
        }
    frontend_config = {
        "server": {
            "port": cfg.get("server", {}).get("port", 5000)
        },
        "llm": {
            "provider": llm_cfg.get("provider", "lm_studio"),
            "base_url": llm_cfg.get("base_url", ""),
            "api_key": _mask_api_key(llm_cfg.get("api_key", "")),
            "model": llm_cfg.get("model", ""),
            "temperature": llm_cfg.get("temperature", 0.3),
            "max_tokens": llm_cfg.get("max_tokens", 4096),
            "timeout": llm_cfg.get("timeout", 300),
            "multimodal": llm_cfg.get("multimodal", True),
            "enable_thinking": llm_cfg.get("enable_thinking", True),
            "reasoning_effort": llm_cfg.get("reasoning_effort")
        },
        "desensitization": {
            "depth": cfg["desensitization"].get("depth", "standard"),
            "placeholder": cfg["desensitization"]["placeholder"],
            "date_format": cfg["desensitization"].get("date_format", "YYYY年MM月DD日")
        },
        "extraction": {
            "auto_merge": cfg.get("extraction", {}).get("auto_merge", False),
            "export_path": cfg.get("extraction", {}).get("export_path", ""),
            "docx_image_extract": cfg.get("extraction", {}).get("docx_image_extract", False),
            "active_template_index": cfg.get("extraction", {}).get("active_template_index", 0),
            "invoice_tolerance": cfg.get("extraction", {}).get("invoice_tolerance", 0.02),
            "invoice_rules": cfg.get("extraction", {}).get("invoice_rules", {"R1": True, "R2": True, "R3": True})
        }
    }
    return jsonify(frontend_config)

@app.route("/api/cleanup_uploads", methods=["POST"])
def api_cleanup_uploads():
    body = request.get_json(force=True) or {}
    keep = body.get("keep", 20)
    removed = cleanup_upload_dir(keep=keep)
    return jsonify({"message": f"已清理 {removed} 个上传缓存文件", "removed": removed})

@app.route("/api/queue/info", methods=["GET"])
def api_queue_info_v2():
    return jsonify(get_queue_info_v2())

@app.route("/api/queue/pause", methods=["POST"])
def api_queue_pause_v2():
    abort_event.set()
    pause_event.set()
    logger.info("队列已暂停，当前任务已中断")
    send_sse("status", {"queue_size": task_queue.qsize(), "processing": False, "paused": True})
    return jsonify({"paused": True, "message": "队列已暂停"})

@app.route("/api/queue/resume", methods=["POST"])
def api_queue_resume_v2():
    abort_event.clear()
    pause_event.clear()
    logger.info("队列已恢复")
    send_sse("status", {"queue_size": task_queue.qsize(), "processing": False, "paused": False})
    return jsonify({"paused": False, "message": "队列已恢复"})

@app.route("/api/queue/tasks/batch-cancel", methods=["POST"])
def api_queue_batch_cancel():
    body = request.get_json(force=True) or {}
    task_ids = body.get("task_ids", [])
    if not task_ids:
        return jsonify({"error": "未指定要取消的任务"}), 400
    from app.shared import CANCELLED_TASK_LOCK, cancelled_task_ids, TaskStatus, abort_event
    from app.task_store import get as ts_get, update_status as ts_update, get_children
    all_ids = list(task_ids)
    for tid in task_ids:
        task = ts_get(tid)
        if task and task.parent_task_id is None and not task.is_chunk:
            children = get_children(tid)
            if children:
                for child in children:
                    if child.status != TaskStatus.COMPLETED:
                        all_ids.append(child.task_id)
    with CANCELLED_TASK_LOCK:
        for tid in all_ids:
            cancelled_task_ids.add(tid)
    abort_event.set()
    for tid in all_ids:
        ts_update(tid, TaskStatus.CANCELLED)
    send_sse("log", {"level": "warning", "message": f"\u5df2\u53d6\u6d88 {len(task_ids)} \u4e2a\u4efb\u52a1"})
    return jsonify({"cancelled": len(task_ids)})

@app.route("/api/queue/tasks/batch-retry", methods=["POST"])
def api_queue_batch_retry():
    body = request.get_json(force=True) or {}
    task_ids = body.get("task_ids", [])
    if not task_ids:
        return jsonify({"error": "未指定要重试的任务"}), 400
    from app.task_store import get as ts_get, update_status as ts_update, get_children
    from app.shared import TaskStatus
    count = 0
    for tid in task_ids:
        task = ts_get(tid)
        if not task:
            continue
        if task.parent_task_id is None and not task.is_chunk:
            children = get_children(tid)
            if children:
                for child in children:
                    if child.status in (TaskStatus.FAILED, TaskStatus.CANCELLED):
                        ts_update(child.task_id, TaskStatus.WAITING)
                        from app.task_store import update_checkpoint
                        update_checkpoint(child.task_id, "", None)
                        with CANCELLED_TASK_LOCK:
                            cancelled_task_ids.discard(child.task_id)
                        enqueue_task_id(child.task_id)
                        count += 1
                ts_update(tid, TaskStatus.WAITING)
            elif task.status in (TaskStatus.FAILED, TaskStatus.CANCELLED):
                ts_update(tid, TaskStatus.WAITING)
                from app.task_store import update_checkpoint as _uc
                _uc(tid, "", None)
                with CANCELLED_TASK_LOCK:
                    cancelled_task_ids.discard(tid)
                enqueue_task_id(tid)
                count += 1
        elif task.status in (TaskStatus.FAILED, TaskStatus.CANCELLED):
            ts_update(tid, TaskStatus.WAITING)
            from app.task_store import update_checkpoint as _uc2
            _uc2(tid, "", None)
            with CANCELLED_TASK_LOCK:
                cancelled_task_ids.discard(tid)
            enqueue_task_id(tid)
            count += 1
    send_sse("log", {"level": "info", "message": f"\u5df2\u91cd\u65b0\u5165\u961f {count} \u4e2a\u4efb\u52a1"})
    return jsonify({"retrying": count})

@app.route("/api/queue/tasks/batch-delete", methods=["POST"])
def api_queue_batch_delete():
    body = request.get_json(force=True) or {}
    task_ids = body.get("task_ids", [])
    if not task_ids:
        return jsonify({"error": "未指定要删除的任务"}), 400
    from app.task_store import delete as ts_delete, get as ts_get, get_children
    from app.shared import TaskStatus
    import os as _os
    # 安全防护：仅允许删除UPLOAD_DIR下的文件
    _safe_root = os.path.abspath(str(shared.UPLOAD_DIR)) + os.sep
    def _safe_unlink(fp):
        if not fp:
            return
        fp_abs = os.path.abspath(fp)
        if os.path.commonpath([fp_abs, _safe_root]) != _safe_root.rstrip(os.sep):
            logger.warning(f"拒绝删除非上传目录文件: {fp}")
            return
        try:
            _os.unlink(fp)
        except OSError:
            pass
    count = 0
    for tid in task_ids:
        task = ts_get(tid)
        if not task:
            continue
        if task.parent_task_id is None and not task.is_chunk:
            children = get_children(tid)
            if children:
                for child in children:
                    _safe_unlink(child.filepath)
                    ts_delete(child.task_id)
                    count += 1
            else:
                _safe_unlink(task.filepath)
            ts_delete(tid)
            with CANCELLED_TASK_LOCK:
                cancelled_task_ids.discard(tid)
        else:
            _safe_unlink(task.filepath)
            ts_delete(tid)
            with CANCELLED_TASK_LOCK:
                cancelled_task_ids.discard(tid)
            count += 1
    send_sse("log", {"level": "info", "message": f"\u5df2\u5f7b\u5e95\u5220\u9664 {count} \u4e2a\u4efb\u52a1"})
    return jsonify({"deleted": count})

@app.route("/api/queue/tasks", methods=["GET"])
def api_queue_tasks():
    mode = request.args.get("mode", "")
    from app.task_store import get_by_mode
    from app.shared import QueueMode, TaskStatus
    try:
        qmode = QueueMode(mode) if mode else None
    except ValueError:
        qmode = None
    tasks = []
    if qmode:
        recent = get_by_mode(qmode, limit=100)
        for t in recent:
            if t.status in (TaskStatus.WAITING, TaskStatus.PROCESSING, TaskStatus.FAILED, TaskStatus.CANCELLED):
                tasks.append(t.to_frontend())
    return jsonify({"tasks": tasks})

@app.route("/api/queue/parent/<parent_id>/merged", methods=["GET"])
def api_queue_parent_merged(parent_id):
    from app.task_store import get_children, get as ts_get
    from app.shared import TaskStatus

    parent = ts_get(parent_id)
    if not parent:
        return jsonify({"error": "任务不存在"}), 404
    children = get_children(parent_id)
    completed = [c for c in children if c.status == TaskStatus.COMPLETED]
    completed.sort(key=lambda c: c.chunk_index)
    merged_text = ""
    merged_original = ""
    merged_replacements = []
    total_duration = 0
    for c in completed:
        if c.result:
            merged_text += c.result.get("text", "")
            merged_original += c.result.get("original_text", c.result.get("text", ""))
            merged_replacements.extend(c.result.get("replacements", []))
            total_duration += c.result.get("duration", 0)
    return jsonify({
        "text": merged_text,
        "original_text": merged_original,
        "replacements": merged_replacements,
        "completed": len(completed),
        "total": len(children),
        "parent_display_name": parent.display_name,
        "total_duration": total_duration,
    })

@app.route("/api/queue/parent/<parent_id>/extract_merged", methods=["GET"])
def api_queue_parent_extract_merged(parent_id):
    from app.task_store import get_children, get as ts_get
    from app.shared import TaskStatus
    parent = ts_get(parent_id)
    if not parent:
        return jsonify({"error": "任务不存在"}), 404
    children = get_children(parent_id)
    completed = [c for c in children if c.status == TaskStatus.COMPLETED and c.result]
    completed.sort(key=lambda c: c.chunk_index)
    merged_items = []
    merged_header = None
    merged_labels = None
    merged_fields = None
    merged_keys = None
    merged_sections = None
    total_len = 0
    for c in completed:
        r = c.result
        total_len += r.get("original_length", 0)
        if merged_header is None:
            merged_header = r.get("extract_header", [])
            merged_labels = r.get("extract_header_labels", [])
            merged_fields = r.get("extract_fields", [])
            merged_keys = r.get("extract_field_keys", [])
            merged_sections = r.get("extract_field_sections", [])
        merged_items.extend(r.get("extract_items", []))
    return jsonify({
        "mode": "extract",
        "header": merged_header,
        "header_labels": merged_labels,
        "fields": merged_fields,
        "field_keys": merged_keys,
        "field_sections": merged_sections,
        "items": merged_items,
        "item_count": len(merged_items),
        "original_length": total_len,
        "completed": len(completed),
        "total": len(children),
        "parent_display_name": parent.display_name,
    })

@app.route("/api/logs", methods=["GET"])
def api_logs():
    last = request.args.get("last", 200, type=int)
    log_path = BASE_DIR / "app.log"
    entries = []
    if log_path.exists():
        with open(log_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        for line in lines[-last:]:
            entries.append({"raw": line.strip()})
    return jsonify({"logs": entries, "count": len(entries)})

